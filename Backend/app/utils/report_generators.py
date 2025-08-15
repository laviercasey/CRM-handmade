import io
import csv
import logging
from datetime import datetime
from typing import Dict, List, Any

logger = logging.getLogger(__name__)

def generate_pdf_report(
    overview: Dict[str, Any],
    status_counts: Dict[str, int],
    activity: List[Dict[str, Any]],
    top_products: List[Dict[str, Any]]
) -> bytes:
    try:
        from fpdf import FPDF
        from fpdf.enums import XPos, YPos
    except ImportError:
        logger.error("PDF export requires fpdf library")
        raise ImportError("PDF export requires fpdf library. Please install it.")
    
    try:
        pdf = FPDF()
        pdf.add_page()
        try:
            pdf.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
            pdf.add_font("DejaVu", "B", "DejaVuSans-Bold.ttf", uni=True)
            font_name = "DejaVu"
        except RuntimeError:
            font_name = "Arial"
        
        pdf.set_font(font_name, "B", 16)
        pdf.cell(0, 10, "Отчет по статистике", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        pdf.ln(10)
        
        pdf.set_font(font_name, "B", 12)
        pdf.cell(0, 10, "Общая статистика:", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font(font_name, "", 10)
        pdf.cell(0, 8, f"Всего заявок: {overview['totalRequests']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(0, 8, f"Всего товаров: {overview['totalProducts']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(0, 8, f"Общая выручка: {overview['totalRevenue']} руб.", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(10)
        
        pdf.set_font(font_name, "B", 12)
        pdf.cell(0, 10, "Заявки по статусам:", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font(font_name, "", 10)
        for status, count in status_counts.items():
            pdf.cell(0, 8, f"{status.capitalize()}: {count}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(10)
        
        pdf.set_font(font_name, "B", 12)
        pdf.cell(0, 10, "Топовые продукты:", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font(font_name, "", 10)
        for product in top_products:
            pdf.cell(0, 8, f"{product['name']} - {product['orderCount']} заказов, {product['totalRevenue']} руб.", 
                  new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        pdf.add_page()
        pdf.set_font(font_name, "B", 12)
        pdf.cell(0, 10, "Активность по дням:", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font(font_name, "", 10)
        
        col_width = pdf.epw / 2
        row_height = 10
        
        pdf.cell(col_width, row_height, "День", border=1, align="C")
        pdf.cell(col_width, row_height, "Количество", border=1, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        for day in activity:
            pdf.cell(col_width, row_height, day["label"], border=1, align="C")
            pdf.cell(col_width, row_height, str(day["count"]), border=1, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        output = io.BytesIO()
        pdf_bytes = pdf.output()
        output.write(pdf_bytes)
        output.seek(0)
        
        return output.getvalue()
    
    except Exception as e:
        logger.error(f"PDF generation error: {str(e)}")
        raise RuntimeError(f"Failed to generate PDF report: {str(e)}")


def generate_csv_report(
    overview: Dict[str, Any],
    status_counts: Dict[str, int],
    activity: List[Dict[str, Any]],
    top_products: List[Dict[str, Any]]
) -> bytes:
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(["Отчет по статистике"])
        writer.writerow([])
        
        writer.writerow(["Общая статистика"])
        writer.writerow(["Всего заявок", overview["totalRequests"]])
        writer.writerow(["Всего товаров", overview["totalProducts"]])
        writer.writerow(["Общая выручка", overview["totalRevenue"]])
        writer.writerow([])
        
        writer.writerow(["Заявки по статусам"])
        for status, count in status_counts.items():
                writer.writerow([status.capitalize(), count])
        writer.writerow([])
        
        writer.writerow(["Активность по дням"])
        writer.writerow(["День", "Количество"])
        for day in activity:
            writer.writerow([day["label"], day["count"]])
        writer.writerow([])
        
        writer.writerow(["Топовые продукты"])
        writer.writerow(["Название", "Количество заказов", "Выручка"])
        for product in top_products:
            writer.writerow([product["name"], product["orderCount"], product["totalRevenue"]])
        
        return output.getvalue().encode("utf-8-sig")
    
    except Exception as e:
        logger.error(f"CSV generation error: {str(e)}")
        raise RuntimeError(f"Failed to generate CSV report: {str(e)}")


def generate_excel_report(
    overview: Dict[str, Any],
    status_counts: Dict[str, int],
    activity: List[Dict[str, Any]],
    top_products: List[Dict[str, Any]]
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        logger.error("Excel export requires openpyxl library")
        raise ImportError("Excel export requires openpyxl library. Please install it.")
    
    try:
        wb = Workbook()
        
        ws_overview = wb.active
        ws_overview.title = "Общая статистика"
        
        ws_overview["A1"] = "Отчет по статистике"
        ws_overview["A1"].font = Font(bold=True, size=14)
        
        ws_overview["A3"] = "Общая статистика"
        ws_overview["A3"].font = Font(bold=True)
        ws_overview["A4"] = "Всего заявок"
        ws_overview["B4"] = overview["totalRequests"]
        ws_overview["A5"] = "Всего товаров"
        ws_overview["B5"] = overview["totalProducts"]
        ws_overview["A6"] = "Общая выручка"
        ws_overview["B6"] = overview["totalRevenue"]
        
        ws_status = wb.create_sheet("Заявки по статусам")
        
        ws_status["A1"] = "Статус"
        ws_status["B1"] = "Количество"
        ws_status["A1"].font = Font(bold=True)
        ws_status["B1"].font = Font(bold=True)
        
        row = 2
        for status, count in status_counts.items():
            ws_status[f"A{row}"] = status.capitalize()
            ws_status[f"B{row}"] = count
            row += 1
        
        ws_activity = wb.create_sheet("Активность")
        
        ws_activity["A1"] = "День"
        ws_activity["B1"] = "Количество"
        ws_activity["A1"].font = Font(bold=True)
        ws_activity["B1"].font = Font(bold=True)
        
        row = 2
        for day in activity:
            ws_activity[f"A{row}"] = day["label"]
            ws_activity[f"B{row}"] = day["count"]
            row += 1
        
        ws_products = wb.create_sheet("Топовые продукты")
        
        ws_products["A1"] = "Название"
        ws_products["B1"] = "Количество заказов"
        ws_products["C1"] = "Выручка"
        ws_products["A1"].font = Font(bold=True)
        ws_products["B1"].font = Font(bold=True)
        ws_products["C1"].font = Font(bold=True)
        
        row = 2
        for product in top_products:
            ws_products[f"A{row}"] = product["name"]
            ws_products[f"B{row}"] = product["orderCount"]
            ws_products[f"C{row}"] = product["totalRevenue"]
            row += 1
        
        for ws in [ws_overview, ws_status, ws_activity, ws_products]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    except Exception as e:
        logger.error(f"Excel generation error: {str(e)}")
        raise RuntimeError(f"Failed to generate Excel report: {str(e)}")
